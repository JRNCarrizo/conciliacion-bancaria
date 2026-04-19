"""
Genera dos .xlsx de ejemplo para import en conciliación (layout por defecto del sistema).
Ejecutar desde la carpeta muestras: py -3 generar_muestras.py
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT_DIR = Path(__file__).resolve().parent

COMPANY_HEADERS = [
    "Fecha\ncontable",
    "División",
    "Tipo",
    "Número",
    "Fecha\nbanco",
    "Debe",
    "Haber",
    "Saldo",
    "Número\ncheque",
    "Observacion",
    "Tipo de\nSujeto",
    "Código\nSujeto",
    "Nombre del\nSujeto",
    "Cuit del\nSujeto",
]


def build_company() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["A1"] = "TES9040-Rut. cons. Resumen bancario (muestra)"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    ws.append(COMPANY_HEADERS)
    for c in range(1, 15):
        ws.cell(row=3, column=c).font = Font(bold=True)

    # (fecha_contable, division, tipo, numero, fecha_banco, debe, haber, saldo, cheque, obs, tipo_suj, cod, nombre, cuit)
    # Importe conciliación = debe − haber (igual al extracto en esta convención).
    data = [
        (
            datetime(2025, 1, 15),
            1,
            "OP",
            2001,
            datetime(2025, 1, 15),
            0,
            25000,
            1_000_000,
            "",
            "Pago Proveedor Demo SA",
            "Proveedor",
            "10",
            "Proveedor Demo",
            "20123456789",
        ),
        (
            datetime(2025, 1, 20),
            1,
            "OP",
            2002,
            datetime(2025, 1, 20),
            0,
            15000,
            985_000,
            "",
            "Débitos varios",
            "Proveedor",
            "11",
            "AFIP / servicios",
            "30701234567",
        ),
        (
            datetime(2025, 1, 22),
            1,
            "RCC",
            3001,
            datetime(2025, 1, 22),
            80000,
            0,
            1_065_000,
            "",
            "Cobranza cliente",
            "Cliente",
            "5",
            "Cliente Ejemplo SRL",
            "30707654321",
        ),
        (
            datetime(2025, 1, 28),
            1,
            "OP",
            2004,
            datetime(2025, 1, 28),
            0,
            5000,
            1_060_000,
            "",
            "Solo en libro (no hay ítem igual en extracto muestra)",
            "Proveedor",
            "99",
            "Varios",
            "20000000000",
        ),
    ]
    for row in data:
        ws.append(list(row))

    out = OUT_DIR / "muestra-plataforma-empresa.xlsx"
    wb.save(out)
    print(f"OK: {out}")


def build_bank() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws["A1"] = "Extracto cuenta corriente — DEMOSTRACIÓN"
    ws["A1"].font = Font(bold=True)
    for _ in range(5):
        ws.append([])
    ws.append(["Período:", datetime(2025, 1, 1), "al", datetime(2025, 1, 31)])
    ws.append(["Fecha", "Código", "Sucursal", "Referencia", "Concepto / detalle", "Observaciones", "Importe"])
    for c in range(1, 8):
        ws.cell(row=8, column=c).font = Font(bold=True)

    movs = [
        (datetime(2025, 1, 15), "DEB", "001", "TRF-001", "Pago Proveedor Demo SA", "Transferencia saliente", -25000),
        (datetime(2025, 1, 20), "VAR", "002", "IMP-445", "Débitos varios", "Impuesto / servicio", -15000),
        (datetime(2025, 1, 22), "CRE", "003", "DEP-CH", "Cobranza cliente", "Acreditación", 80000),
        (datetime(2025, 1, 25), "COM", "099", "COM-MES", "Comisión bancaria", "Solo en banco (sin par en libro muestra)", -999),
    ]
    for m in movs:
        ws.append(list(m))

    out = OUT_DIR / "muestra-extracto-banco.xlsx"
    wb.save(out)
    print(f"OK: {out}")


if __name__ == "__main__":
    build_company()
    build_bank()
